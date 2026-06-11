import base64
import json
import unicodedata
from datetime import datetime
from pathlib import Path


import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(
    page_title="전국 특산물 지도",
    page_icon="🗺️",
    layout="wide",
)

BASE_DIR = Path(__file__).resolve().parent
PRODUCT_INFO_PATH = BASE_DIR / "product_info.json"
ERROR_REQUEST_PATH = BASE_DIR / "오류수정요청.xlsx"
ICON_DIR_CANDIDATES = [BASE_DIR / "icons", BASE_DIR / "아이콘"]



def save_error_request(payload: dict) -> tuple[bool, str]:
    """오류수정요청.xlsx에 수정 요청 로그를 시간순으로 누적 저장합니다."""
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        return False, f"openpyxl 라이브러리를 불러오지 못했습니다: {exc}"

    headers = ["접수시간", "권역", "지역", "특산물", "수정내용", "요청ID"]
    try:
        if ERROR_REQUEST_PATH.exists():
            wb = load_workbook(ERROR_REQUEST_PATH)
            ws = wb.active
            if ws.max_row < 1 or [cell.value for cell in ws[1][:len(headers)]] != headers:
                ws.insert_rows(1)
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "오류수정요청"
            ws.append(headers)

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(payload.get("region", "")).strip(),
            str(payload.get("city", "")).strip(),
            str(payload.get("product", "")).strip(),
            str(payload.get("requestText", "")).strip(),
            str(payload.get("requestId", "")).strip(),
        ])

        widths = {"A": 21, "B": 14, "C": 16, "D": 18, "E": 46, "F": 26}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        wb.save(ERROR_REQUEST_PATH)
        return True, f"오류수정요청.xlsx 저장 완료: {ERROR_REQUEST_PATH}"
    except Exception as exc:
        return False, f"오류수정요청.xlsx 저장 실패: {exc}"



def load_product_icons() -> dict:
    """icons/ 또는 아이콘/ 폴더의 PNG 파일을 HTML에서 바로 표시할 수 있도록 Base64 data URI로 변환합니다."""
    icons = {}
    icon_dir = next((path for path in ICON_DIR_CANDIDATES if path.exists() and path.is_dir()), None)
    if icon_dir is None:
        st.info("icons 또는 아이콘 폴더가 없습니다. 특산물 아이콘 없이 텍스트만 표시합니다.")
        return icons

    for file in icon_dir.rglob("*.png"):
        if file.name.startswith("._"):
            continue
        try:
            key = unicodedata.normalize("NFC", file.stem.strip())
            encoded = base64.b64encode(file.read_bytes()).decode("ascii")
            icons[key] = f"data:image/png;base64,{encoded}"
        except Exception as exc:
            st.warning(f"아이콘 파일을 불러오지 못했습니다: {file.name} / {exc}")
    # 파일명이 product_info.json의 특산물명과 조금 다른 경우를 위한 최소 별칭입니다.
    if "송이버섯" not in icons and "송이" in icons:
        icons["송이버섯"] = icons["송이"]
    return icons


def load_product_info() -> dict:
    """product_info.json을 읽어서 HTML 내부 JS에서 사용할 딕셔너리로 반환합니다."""
    if not PRODUCT_INFO_PATH.exists():
        st.warning("product_info.json 파일이 없습니다. 특산물 상세정보 없이 기본 지도만 표시합니다.")
        return {}

    try:
        return json.loads(PRODUCT_INFO_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        st.error(f"product_info.json 형식 오류: {exc}")
        return {}


HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>전국 특산물 권역 지도</title>
<script src="https://cdn.jsdelivr.net/npm/d3@7"></script>
<style>
*{box-sizing:border-box} body{margin:0;font-family:'Pretendard','Noto Sans KR',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;color:#111827}
.wrap{max-width:1180px;margin:0 auto;padding:24px 16px 30px} h1{margin:0 0 8px;text-align:center;font-size:32px} .sub{text-align:center;color:#5b6470;margin:0 0 18px}
.legend{display:flex;flex-wrap:wrap;gap:10px 16px;justify-content:center;background:#fff;border:1px solid #e5e7eb;border-radius:14px;padding:11px;margin-bottom:16px}
.leg{display:flex;align-items:center;gap:7px;font-size:13px;font-weight:700;cursor:pointer} .dot{width:13px;height:13px;border-radius:999px;border:1px solid rgba(0,0,0,.15)}
.grid{display:grid;grid-template-columns:minmax(0,1fr) 340px;gap:18px;align-items:start} .mapbox{background:#ffffff;border:1px solid #e5e7eb;border-radius:16px;padding:14px;min-height:720px;position:relative}
#map{width:100%;height:auto;display:block} .province{stroke:#fff;stroke-width:1.4;stroke-linejoin:round;stroke-linecap:round;cursor:pointer;transition:filter .12s, opacity .12s, fill .12s;vector-effect:non-scaling-stroke} .province:hover{filter:brightness(1.05) drop-shadow(0 2px 3px rgba(0,0,0,.18))} .province.active{stroke:#fff;stroke-width:1.4;filter:drop-shadow(0 2px 5px rgba(0,0,0,.22))} .selected-outline{fill:none;stroke:#111;stroke-width:1.8;stroke-linejoin:round;stroke-linecap:round;pointer-events:none;vector-effect:non-scaling-stroke;filter:drop-shadow(0 2px 5px rgba(0,0,0,.22))} .ulleung.active .province{stroke:#111;stroke-width:1.8;filter:drop-shadow(0 2px 5px rgba(0,0,0,.22))}
.label-bg{fill:rgba(255,255,255,.85);stroke:#e5e7eb;stroke-width:1;filter:none} .label{font-size:13px;font-weight:700;text-anchor:middle;dominant-baseline:middle;pointer-events:none;paint-order:stroke;stroke:#fff;stroke-width:3px;stroke-linejoin:round}
.inset-title{font-size:13px;font-weight:900;text-anchor:middle} .inset-box{fill:#fff;stroke:#9ca3af;stroke-dasharray:4 3;rx:10}
.panel{background:#fff;border:1px solid #e5e7eb;border-radius:18px;padding:16px;position:sticky;top:12px;max-height:calc(100vh - 24px);overflow:auto} .panel h2{margin:0 0 14px;text-align:center;font-size:20px}
.selected{border:1px solid #e5e7eb;border-radius:14px;overflow:hidden} .head{display:flex;justify-content:space-between;align-items:center;padding:12px 14px;font-weight:900;border-bottom:1px solid #e5e7eb} .count{font-size:13px;color:#6b7280}
.list{padding:8px 10px 12px} .row{display:grid;grid-template-columns:88px 1fr;gap:8px;padding:7px 4px;border-bottom:1px dashed #e5e7eb;font-size:14px;align-items:center}.row:last-child{border-bottom:0} .city{font-weight:800} .prod{line-height:1.45;color:#374151;display:flex;flex-wrap:wrap;gap:5px}
.product-btn{appearance:none;border:1px solid #d1d5db;background:#fff;border-radius:999px;padding:4px 9px 4px 6px;margin:2px 4px 2px 0;font-size:13px;font-weight:800;color:#374151;cursor:pointer;transition:.12s;display:inline-flex;align-items:center;gap:6px;min-height:34px}.product-icon{width:24px;height:24px;object-fit:contain;flex:0 0 auto}.product-icon.missing{display:none}.product-btn:hover{transform:translateY(-1px);box-shadow:0 2px 8px rgba(0,0,0,.12);border-color:#9ca3af}.product-btn.active{color:#111;border-color:#111;background:#f9fafb}.detail-card{margin:12px 10px 14px;padding:13px;border:1px solid #e5e7eb;border-radius:14px;background:#f9fafb}.detail-title{font-size:18px;font-weight:900;margin:0 0 8px}.detail-meta{font-size:13px;color:#4b5563;line-height:1.7}.detail-actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}.detail-actions a,.detail-actions button{display:inline-flex;text-decoration:none;color:#111;border:1px solid #d1d5db;background:#fff;border-radius:9px;padding:7px 10px;font-size:12px;font-weight:800;cursor:pointer}.detail-actions a:hover,.detail-actions button:hover{background:#f3f4f6}.error-btn{color:#991b1b!important;border-color:#fecaca!important}.modal-backdrop{position:fixed;inset:0;background:rgba(15,23,42,.42);display:none;align-items:center;justify-content:center;z-index:100}.modal{width:min(420px,calc(100vw - 28px));background:#fff;border-radius:16px;border:1px solid #e5e7eb;box-shadow:0 18px 55px rgba(0,0,0,.22);padding:16px}.modal h3{margin:0 0 8px;font-size:18px}.modal p{margin:0 0 10px;color:#6b7280;font-size:13px;line-height:1.55}.modal textarea{width:100%;min-height:130px;border:1px solid #d1d5db;border-radius:12px;padding:10px;font-family:inherit;font-size:14px;resize:vertical}.modal-actions{display:flex;justify-content:flex-end;gap:8px;margin-top:10px}.modal-actions button{border:1px solid #d1d5db;background:#fff;border-radius:9px;padding:8px 12px;font-weight:800;cursor:pointer}.modal-actions .submit{background:#111827;color:#fff;border-color:#111827}.toast{position:fixed;right:18px;bottom:18px;background:#111827;color:#fff;border-radius:10px;padding:10px 12px;font-size:13px;display:none;z-index:120}.hint{color:#6b7280;text-align:center;line-height:1.7;padding:28px 8px} .err{color:#b91c1c;background:#fff1f2;border:1px solid #fecdd3;border-radius:12px;padding:12px;line-height:1.6}
.tip{position:fixed;display:none;z-index:50;background:rgba(17,24,39,.92);color:white;border-radius:9px;padding:8px 10px;font-size:12px;pointer-events:none;line-height:1.5;max-width:250px}
.note{margin-top:16px;background:#fff;border:1px solid #e5e7eb;border-radius:14px;padding:12px;text-align:center;color:#6b7280;font-size:13px;line-height:1.6}
@media(max-width:900px){.grid{grid-template-columns:1fr}.panel{position:static;max-height:none}.mapbox{min-height:0}h1{font-size:26px}}
</style>
</head>
<body><div class="wrap"><h1>전국 특산물 지도</h1><p class="sub">권역을 클릭하면 엑셀에 입력된 해당 시/군구 특산물 리스트가 표시됩니다.</p><div class="legend" id="legend"></div><div class="grid"><main class="mapbox"><svg id="map" viewBox="0 0 650 780" role="img" aria-label="대한민국 권역별 지도"></svg></main><aside class="panel"><h2>권역을 선택해주세요</h2><div id="sidebar" class="hint">지도에서 원하는 권역을 클릭하세요.</div></aside></div><div class="note">※ 데이터는 제공된 엑셀의 권역 / 시군구 / 특산물정보만 사용했습니다. 지도 경계는 공개 행정구역 GeoJSON을 불러와 표시하며, 울릉도·독도 inset은 별도 좌표 기반 경계로 표시합니다.</div></div><div class="tip" id="tip"></div><div class="modal-backdrop" id="errorModal"><div class="modal"><h3>오류수정요청</h3><p id="errorModalMeta">수정할 내용을 입력해주세요.</p><textarea id="errorRequestText" placeholder="예: 소비기한을 3일로 수정해주세요."></textarea><div class="modal-actions"><button type="button" id="errorCancelBtn">취소</button><button type="button" class="submit" id="errorSubmitBtn">제출</button></div></div></div><div class="toast" id="toast"></div>
<script>
const PRODUCT_INFO=__PRODUCT_INFO_JSON__;
const PRODUCT_ICONS=__PRODUCT_ICONS_JSON__;
const REGION_DATA={"서울경기": {"color": "#ef9a9a", "provinceNames": ["서울특별시", "경기도", "인천광역시", "세종특별자치시"], "cities": [{"name": "양주시", "p": "부추"}, {"name": "가평시", "p": "잣"}, {"name": "파주시", "p": "장단콩"}, {"name": "이천시", "p": "이천쌀"}, {"name": "안성시", "p": "안성배"}]}, "강원도": {"color": "#a5d6a7", "provinceNames": ["강원도", "강원특별자치도"], "cities": [{"name": "춘천시", "p": "닭고기"}, {"name": "원주시", "p": "복숭아"}, {"name": "강릉시", "p": "초당콩"}, {"name": "동해시", "p": "문어"}, {"name": "태백시", "p": "고랭지배추"}, {"name": "속초시", "p": "붉은대게"}, {"name": "삼척시", "p": "장뇌삼,봄굴"}, {"name": "평창군", "p": "황태,메밀"}]}, "충청남도": {"color": "#ffe082", "provinceNames": ["충청남도", "대전광역시"], "cities": [{"name": "천안시", "p": "호두"}, {"name": "공주시", "p": "밤"}, {"name": "보령시", "p": "대하"}, {"name": "논산시", "p": "딸기"}, {"name": "서천시", "p": "전어"}]}, "충청북도": {"color": "#c5e1a5", "provinceNames": ["충청북도"], "cities": [{"name": "제천시", "p": "황기"}, {"name": "보은군", "p": "대추"}, {"name": "영동군", "p": "포도"}, {"name": "증평군", "p": "인삼"}, {"name": "괴산군", "p": "대학찰옥수수"}]}, "전라남도": {"color": "#ce93d8", "provinceNames": ["전라남도"], "cities": [{"name": "목포시", "p": "세발낙지,홍어"}, {"name": "여수시", "p": "돌산갓,멸치"}, {"name": "나주시", "p": "배"}, {"name": "광양시", "p": "매실"}, {"name": "담양군", "p": "죽순"}, {"name": "곡성군", "p": "토란"}, {"name": "구례군", "p": "산수유"}, {"name": "고흥군", "p": "유자"}, {"name": "보성군", "p": "녹차"}, {"name": "해남군", "p": "고구마,배추"}, {"name": "영암군", "p": "무화과"}, {"name": "무안군", "p": "양파"}, {"name": "영광군", "p": "굴비"}, {"name": "완도군", "p": "전복,김"}, {"name": "진도군", "p": "대파"}, {"name": "신안군", "p": "천일염"}]}, "전라북도": {"color": "#ffcc80", "provinceNames": ["전라북도", "전북특별자치도", "광주광역시"], "cities": [{"name": "전주시", "p": "콩나물"}, {"name": "군산시", "p": "박대"}, {"name": "익산시", "p": "서동마"}, {"name": "남원시", "p": "미꾸라지"}, {"name": "완주군", "p": "생강"}, {"name": "진안군", "p": "홍삼대제"}, {"name": "무주군", "p": "천마"}, {"name": "임실군", "p": "치즈"}, {"name": "순창군", "p": "고추장"}, {"name": "고창군", "p": "복분자,풍천장어"}, {"name": "부안군", "p": "오디"}]}, "경상북도": {"color": "#90caf9", "provinceNames": ["경상북도", "대구광역시"], "cities": [{"name": "포항시", "p": "과메기"}, {"name": "김천시", "p": "자두"}, {"name": "안동시", "p": "소주"}, {"name": "상주시", "p": "감"}, {"name": "문경시", "p": "오미자"}, {"name": "청송군", "p": "사과"}, {"name": "의성군", "p": "마늘"}, {"name": "성주군", "p": "참외"}, {"name": "봉화군", "p": "송이버섯"}]}, "경상남도": {"color": "#80cbc4", "provinceNames": ["경상남도", "부산광역시", "울산광역시"], "cities": [{"name": "창원시", "p": "미더덕"}, {"name": "통영시", "p": "굴"}, {"name": "함양군", "p": "산양삼"}, {"name": "남해군", "p": "시금치"}, {"name": "함안군", "p": "수박"}, {"name": "고성군", "p": "방울토마토"}]}, "제주도": {"color": "#b39ddb", "provinceNames": ["제주특별자치도"], "cities": [{"name": "서귀포시", "p": "한라봉,흑돼지"}]}, "울릉도": {"color": "#b0bec5", "provinceNames": [], "cities": [{"name": "울릉군", "p": "명이나물,오징어"}]}};
const REGION_ORDER=["서울경기", "강원도", "충청남도", "충청북도", "전라남도", "전라북도", "경상북도", "경상남도", "제주도", "울릉도"];
const GEO_URLS=[
 'https://raw.githubusercontent.com/southkorea/southkorea-maps/master/kostat/2018/json/skorea-provinces-2018-geo.json',
 'https://raw.githubusercontent.com/southkorea/southkorea-maps/master/kostat/2013/json/skorea_provinces_geo_simple.json'
];
const ALIASES={'서울':'서울특별시','부산':'부산광역시','대구':'대구광역시','인천':'인천광역시','광주':'광주광역시','대전':'대전광역시','울산':'울산광역시','세종':'세종특별자치시','경기':'경기도','강원':'강원도','충북':'충청북도','충남':'충청남도','전북':'전라북도','전남':'전라남도','경북':'경상북도','경남':'경상남도','제주':'제주특별자치도'};
const provinceToRegion={}; Object.entries(REGION_DATA).forEach(([r,d])=>d.provinceNames.forEach(n=>provinceToRegion[n]=r));
function cleanName(v){if(!v) return ''; let s=String(v).trim(); return ALIASES[s]||s;}
function getProvinceName(p){const keys=['CTP_KOR_NM','CTPRVN_NM','SIDO_NM','NAME_1','name','Name','NAME','adm_nm']; for(const k of keys) if(p&&p[k]) return cleanName(p[k]); return '';}
function regionForFeature(f){let n=getProvinceName(f.properties); if(provinceToRegion[n]) return provinceToRegion[n]; const vals=Object.values(f.properties||{}).map(cleanName); for(const v of vals) if(provinceToRegion[v]) return provinceToRegion[v]; return null;}
const svg=d3.select('#map'), W=650, H=780, tip=document.getElementById('tip'), sidebar=document.getElementById('sidebar');
let mapFeatures=[], mapPath=null;
function esc(s){return String(s).replace(/[&<>'"]/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[m]));}
function splitProducts(p){return String(p).split(/[,·]/).map(x=>x.trim()).filter(Boolean);}
function normKey(s){return String(s||'').normalize ? String(s||'').normalize('NFC') : String(s||'');}
function iconTag(product){
 const src=PRODUCT_ICONS[normKey(product)] || '';
 return src ? `<img class="product-icon" src="${src}" alt="${esc(product)} 아이콘">` : `<span class="product-icon missing"></span>`;
}
function productButton(region, city, product){
 return `<button type="button" class="product-btn" data-region="${esc(region)}" data-city="${esc(city)}" data-product="${esc(product)}">${iconTag(product)}<span>${esc(product)}</span></button>`;
}
let currentErrorTarget=null;
const errorModal=document.getElementById('errorModal');
const errorRequestText=document.getElementById('errorRequestText');
const errorModalMeta=document.getElementById('errorModalMeta');
const toast=document.getElementById('toast');
function notify(msg){toast.textContent=msg;toast.style.display='block';setTimeout(()=>{toast.style.display='none';},2400);}
function postStreamlitValue(value){
  window.parent.postMessage({isStreamlitMessage:true,type:'streamlit:setComponentValue',value:value,dataType:'json'}, '*');
}
window.parent.postMessage({isStreamlitMessage:true,type:'streamlit:componentReady',apiVersion:1}, '*');
function openErrorModal(region, city, product){
  currentErrorTarget={region,city,product};
  errorModalMeta.innerHTML=`<b>${esc(region)} / ${esc(city)} / ${esc(product)}</b><br>수정 요청 내용을 입력하면 오류수정요청.xlsx에 시간별로 누적 저장됩니다.`;
  errorRequestText.value='';
  errorModal.style.display='flex';
  setTimeout(()=>errorRequestText.focus(),50);
}
function closeErrorModal(){errorModal.style.display='none';currentErrorTarget=null;}
document.getElementById('errorCancelBtn').addEventListener('click',closeErrorModal);
errorModal.addEventListener('click',(event)=>{if(event.target===errorModal) closeErrorModal();});
document.getElementById('errorSubmitBtn').addEventListener('click',()=>{
  const requestText=errorRequestText.value.trim();
  if(!requestText){notify('수정내용을 입력해주세요.');return;}
  if(!currentErrorTarget){notify('특산물을 다시 선택해주세요.');return;}
  const payload={type:'error_request',requestId:`${Date.now()}-${Math.random().toString(16).slice(2)}`,requestText,...currentErrorTarget};
  postStreamlitValue(payload);
  closeErrorModal();
  notify('오류수정요청이 접수되었습니다.');
});
function showProductDetail(region, city, product){
 const box=document.getElementById('productDetail');
 if(!box) return;
 document.querySelectorAll('.product-btn').forEach(b=>b.classList.toggle('active', b.dataset.region===region && b.dataset.city===city && b.dataset.product===product));
 const info = (((PRODUCT_INFO||{})[region]||{})[city]||{})[product] || {};
 const shelf = info['소비기한'] || 'JSON 파일에 소비기한을 입력해주세요.';
 const storage = info['보관방법'] || 'JSON 파일에 보관방법을 입력해주세요.';
 const recipeLink = info['레시피링크'] || '';
 const recipe=encodeURIComponent(`${product} 레시피`);
 const recipeButton = recipeLink ? `<a href="${esc(recipeLink)}" target="_blank" rel="noopener">등록 레시피</a>` : `<a href="https://www.google.com/search?q=${recipe}" target="_blank" rel="noopener">레시피 검색</a>`;
 box.innerHTML=`<div class="detail-title">${esc(product)}</div><div class="detail-meta"><b>권역</b> ${esc(region)}<br><b>지역</b> ${esc(city)}<br><b>소비기한</b> ${esc(shelf)}<br><b>보관방법</b> ${esc(storage)}</div><div class="detail-actions">${recipeButton}<button type="button" class="error-btn" onclick="openErrorModal('${esc(region)}','${esc(city)}','${esc(product)}')">오류수정요청</button></div>`;
}
function updateSelectedOutline(r){
 d3.selectAll('.selected-outline-layer').remove();
 if(!mapFeatures.length || !mapPath || r==='울릉도') return;
 const selectedFeatures = mapFeatures.filter(f=>regionForFeature(f)===r);
 if(!selectedFeatures.length) return;
 svg.append('g')
   .attr('class','selected-outline-layer')
   .selectAll('path')
   .data(selectedFeatures)
   .join('path')
   .attr('class','selected-outline')
   .attr('d',mapPath);
}
function showRegion(r){
 const d=REGION_DATA[r];
 d3.selectAll('.province').classed('active',x=>regionForFeature(x)===r);
 d3.selectAll('.ulleung').classed('active',r==='울릉도');
 updateSelectedOutline(r);
 sidebar.className='selected';
 sidebar.innerHTML=`<div class="head" style="background:${d.color}22"><span>${r}</span><span class="count">${d.cities.length}개 지역</span></div><div id="productDetail" class="detail-card"><div class="detail-title">특산물을 선택하세요</div><div class="detail-meta">아래 특산물 버튼을 누르면 상세 정보와 레시피/오류수정요청 버튼이 표시됩니다.</div></div><div class="list">${d.cities.map(c=>`<div class="row"><span class="city">${esc(c.name)}</span><span class="prod">${splitProducts(c.p).map(prod=>productButton(r,c.name,prod)).join('')}</span></div>`).join('')}</div>`;
 sidebar.querySelectorAll('.product-btn').forEach(btn=>btn.addEventListener('click',()=>showProductDetail(btn.dataset.region, btn.dataset.city, btn.dataset.product)));
}
function addLegend(){const legend=document.getElementById('legend'); REGION_ORDER.forEach(r=>{const d=REGION_DATA[r]; const el=document.createElement('div'); el.className='leg'; el.innerHTML=`<span class="dot" style="background:${d.color}"></span><span>${r}</span>`; el.onclick=()=>showRegion(r); legend.appendChild(el);});}
async function loadGeo(){let last; for(const url of GEO_URLS){try{const res=await fetch(url); if(!res.ok) throw new Error(res.status); return await res.json();}catch(e){last=e;}} throw last||new Error('지도 데이터를 불러오지 못했습니다.');}
function draw(geo){
 const features=geo.features.filter(f=>regionForFeature(f));
 const projection=d3.geoMercator().fitExtent([[35,25],[595,735]], {type:'FeatureCollection',features});
 const path=d3.geoPath(projection);
 mapFeatures=features;
 mapPath=path;
 svg.append('rect').attr('x',0).attr('y',0).attr('width',W).attr('height',H).attr('fill','transparent');
 svg.selectAll('path.province').data(features).join('path').attr('class','province').attr('d',path).attr('fill',d=>REGION_DATA[regionForFeature(d)].color).attr('opacity',.88)
 .on('mousemove',(event,d)=>{const r=regionForFeature(d), rows=REGION_DATA[r].cities.slice(0,4).map(c=>`${c.name} - ${c.p}`).join('<br>'); tip.style.display='block'; tip.style.left=event.clientX+12+'px'; tip.style.top=event.clientY-28+'px'; tip.innerHTML=`<b>${r}</b><br>${rows}`;})
 .on('mouseleave',()=>tip.style.display='none').on('click',(event,d)=>showRegion(regionForFeature(d)));
 // 권역명 위치: 각 권역의 실제 SVG 경계(bounds)를 계산해서 그 구역 안쪽에 배치합니다.
 // 고정 좌표가 아니라 지도 투영 결과를 기준으로 잡기 때문에 화면 크기/지도 비율이 달라져도 권역 위에 유지됩니다.
 const labelAnchorRatio={
   '서울경기':[0.50,0.47],
   '강원도':[0.56,0.43],
   '충청남도':[0.43,0.50],
   '충청북도':[0.53,0.51],
   '전라북도':[0.48,0.50],
   '전라남도':[0.45,0.55],
   '경상북도':[0.57,0.49],
   '경상남도':[0.55,0.53]
 };
 const labelWidth={
   '서울경기':104,'강원도':88,'충청남도':104,'충청북도':104,
   '전라북도':104,'전라남도':104,'경상북도':104,'경상남도':104
 };
 const labelNudge={
   '서울경기':[0,0],
   '강원도':[0,0],
   '충청남도':[0,0],
   '충청북도':[0,0],
   '전라북도':[0,0],
   '전라남도':[0,0],
   '경상북도':[0,0],
   '경상남도':[0,0]
 };
 function featureCollectionForRegion(r){
   return {type:'FeatureCollection',features:features.filter(f=>regionForFeature(f)===r)};
 }
 function anchorByBounds(r){
   const fc=featureCollectionForRegion(r);
   if(!fc.features.length) return null;
   const nudge=labelNudge[r]||[0,0];

   // 경상북도는 도서 지역까지 포함한 bounds 중심을 쓰면 글자탭이 본토 밖으로 밀릴 수 있습니다.
   // 면적 기반 centroid를 사용해 실제 경상북도 본토 영역 위에 라벨을 고정합니다.
   if(r==='경상북도'){
     const c=path.centroid(fc);
     return [c[0]+nudge[0], c[1]+nudge[1]];
   }

   const b=path.bounds(fc);
   const ratio=labelAnchorRatio[r]||[0.5,0.5];
   const x=b[0][0]+(b[1][0]-b[0][0])*ratio[0]+nudge[0];
   const y=b[0][1]+(b[1][1]-b[0][1])*ratio[1]+nudge[1];
   return [x,y];
 }
 function drawRegionLabel(r,x,y){
   const w=labelWidth[r]||96;
   const g=svg.append('g').attr('class','region-name').style('pointer-events','none');
   g.append('rect').attr('class','label-bg').attr('x',x-w/2).attr('y',y-17).attr('width',w).attr('height',34).attr('rx',17);
   g.append('text').attr('class','label').attr('x',x).attr('y',y).attr('fill',REGION_DATA[r].color).text(r);
 }
 ['서울경기','강원도','충청남도','충청북도','전라북도','전라남도','경상북도','경상남도']
   .forEach(r=>{const p=anchorByBounds(r); if(p) drawRegionLabel(r,p[0],p[1]);});
 // 제주 inset
 const jejuFeature=features.find(f=>regionForFeature(f)==='제주도'); if(jejuFeature){const p2=d3.geoMercator().fitExtent([[60,675],[230,750]], jejuFeature); const path2=d3.geoPath(p2); svg.append('rect').attr('class','inset-box').attr('x',45).attr('y',660).attr('width',205).attr('height',95); svg.append('path').datum(jejuFeature).attr('class','province').attr('d',path2).attr('fill',REGION_DATA['제주도'].color).attr('opacity',.88).on('click',()=>showRegion('제주도')); svg.append('text').attr('class','label').attr('x',147).attr('y',715).attr('fill',REGION_DATA['제주도'].color).text('제주도');}
 // 울릉도·독도 inset: 기존 임의 타원/점 표현을 제거하고, 좌표 기반 GeoJSON 형태로 표시합니다.
 // 주 지도 GeoJSON에는 울릉군/독도 세부 경계가 안정적으로 포함되지 않아, inset 전용 경계 좌표를 별도로 사용합니다.
 const ULLEUNG_DOKDO_GEO={
   type:'FeatureCollection',
   features:[
     {type:'Feature',properties:{name:'울릉도',kind:'island'},geometry:{type:'Polygon',coordinates:[[[130.817,37.545],[130.836,37.538],[130.858,37.518],[130.873,37.491],[130.866,37.472],[130.843,37.458],[130.813,37.456],[130.789,37.469],[130.771,37.491],[130.768,37.516],[130.787,37.537],[130.817,37.545]]]}},
     {type:'Feature',properties:{name:'독도 동도',kind:'dokdo'},geometry:{type:'Polygon',coordinates:[[[131.870,37.243],[131.873,37.242],[131.875,37.240],[131.874,37.238],[131.871,37.237],[131.868,37.238],[131.867,37.241],[131.870,37.243]]]}},
     {type:'Feature',properties:{name:'독도 서도',kind:'dokdo'},geometry:{type:'Polygon',coordinates:[[[131.864,37.244],[131.867,37.243],[131.868,37.241],[131.867,37.239],[131.864,37.238],[131.861,37.239],[131.860,37.242],[131.864,37.244]]]}}
   ]
 };
 const ug=svg.append('g').attr('class','ulleung').style('cursor','pointer')
   .on('click',()=>showRegion('울릉도'))
   .on('mousemove',(event)=>{tip.style.display='block';tip.style.left=event.clientX+12+'px';tip.style.top=event.clientY-28+'px';tip.innerHTML='<b>울릉도·독도</b><br>울릉군 - '+REGION_DATA['울릉도'].cities[0].p;})
   .on('mouseleave',()=>tip.style.display='none');
 ug.append('rect').attr('class','inset-box').attr('x',500).attr('y',65).attr('width',130).attr('height',115);
 const islandProjection=d3.geoMercator().fitExtent([[512,88],[620,154]], ULLEUNG_DOKDO_GEO);
 const islandPath=d3.geoPath(islandProjection);
 ug.selectAll('path.island').data(ULLEUNG_DOKDO_GEO.features).join('path')
   .attr('class','province island')
   .attr('d',islandPath)
   .attr('fill',REGION_DATA['울릉도'].color)
   .attr('opacity',d=>d.properties.kind==='dokdo'?.9:.78)
   .attr('stroke','#475569');
 ug.append('text').attr('class','inset-title').attr('x',558).attr('y',84).text('울릉도·독도');
 ug.append('text').attr('class','inset-title').attr('x',548).attr('y',166).style('font-size','11px').text('울릉도');
 ug.append('text').attr('class','inset-title').attr('x',604).attr('y',166).style('font-size','11px').text('독도');
}
addLegend(); loadGeo().then(draw).catch(e=>{document.querySelector('.mapbox').innerHTML='<div class="err"><b>지도 데이터를 불러오지 못했습니다.</b><br>인터넷 연결 상태에서 HTML을 다시 열어주세요.<br>'+e.message+'</div>';});
</script></body></html>
"""

product_info = load_product_info()
product_icons = load_product_icons()
html_code = (
    HTML_TEMPLATE
    .replace("__PRODUCT_INFO_JSON__", json.dumps(product_info, ensure_ascii=False))
    .replace("__PRODUCT_ICONS_JSON__", json.dumps(product_icons, ensure_ascii=False))
)

component_value = components.html(html_code, height=930, scrolling=True)

if isinstance(component_value, dict) and component_value.get("type") == "error_request":
    request_id = component_value.get("requestId")
    if request_id and st.session_state.get("last_error_request_id") != request_id:
        st.session_state["last_error_request_id"] = request_id
        ok, msg = save_error_request(component_value)
        if ok:
            st.success(msg)
        else:
            st.error(msg)

if ERROR_REQUEST_PATH.exists():
    with ERROR_REQUEST_PATH.open("rb") as fp:
        st.download_button(
            "오류수정요청.xlsx 다운로드",
            data=fp,
            file_name="오류수정요청.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

